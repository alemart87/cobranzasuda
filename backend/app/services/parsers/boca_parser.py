"""Parse Boca de Cobranzas XLSX into a list of dicts."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


REQUIRED_COLUMNS = ["Descripción", "Importe", "Cobrador", "Fecha Pago", "Cod. Concepto"]


def parse_boca(path: str | Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(f"Boca: faltan columnas requeridas: {missing}")

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if row.get("Descripción"):
            rows.append(row)
    return rows
