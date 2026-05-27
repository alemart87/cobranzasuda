"""Parse Cobrado 186 Voicenter XLSX. Columns by position because headers duplicate."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


def parse_cobrado(path: str | Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Report" in wb.sheetnames:
        ws = wb["Report"]
    else:
        ws = wb[wb.sheetnames[0]]

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        descripcion = ws.cell(r, 3).value
        if not descripcion:
            continue
        rows.append({
            "Concepto": ws.cell(r, 1).value,
            "Cuota": ws.cell(r, 2).value,
            "Descripción": descripcion,
            "Recibo": ws.cell(r, 4).value,
            "F/Pago": ws.cell(r, 5).value,
            "Importe": ws.cell(r, 6).value,
            "Total": ws.cell(r, 8).value,
            "Cobrador": ws.cell(r, 10).value,
            "Observación": ws.cell(r, 11).value,
        })
    return rows
