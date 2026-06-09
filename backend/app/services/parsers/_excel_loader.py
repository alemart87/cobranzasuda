"""Universal Excel loader — soporta tanto .xls (xlrd) como .xlsx (openpyxl).

Devuelve una estructura uniforme:
    {sheet_name: [[v00, v01, ...], [v10, v11, ...], ...]}

Las fechas de .xls (que xlrd devuelve como float) se convierten a `datetime`.
"""
from __future__ import annotations

from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any


def _is_xls(path: Path) -> bool:
    """Detecta por magic bytes — más confiable que la extensión."""
    try:
        with open(path, "rb") as f:
            head = f.read(8)
        # OLE2 signature de .xls
        if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            return True
        # ZIP signature de .xlsx
        if head[:4] == b"PK\x03\x04":
            return False
    except Exception:
        pass
    return str(path).lower().endswith(".xls")


def _max_rows() -> int:
    """Tope de filas por hoja (insurance contra archivos absurdos). 0 = sin tope."""
    try:
        from ...core.config import settings
        return int(settings.upload_max_rows) or 0
    except Exception:
        return 300000


def _load_xlsx(path: Path) -> dict[str, list[list[Any]]]:
    import openpyxl
    # read_only=True → lectura en streaming (memoria acotada); sin él, openpyxl
    # carga TODO el workbook en RAM y un archivo grande revienta el proceso.
    cap = _max_rows()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: dict[str, list[list[Any]]] = {}
    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
                if cap and len(rows) >= cap:
                    raise ValueError(
                        f"La hoja '{sname}' supera el máximo de {cap} filas. "
                        f"El archivo parece demasiado grande o tener formato incorrecto."
                    )
            out[sname] = rows
    finally:
        wb.close()
    return out


def _load_xls(path: Path) -> dict[str, list[list[Any]]]:
    import xlrd
    from xlrd import xldate_as_tuple

    cap = _max_rows()
    book = xlrd.open_workbook(str(path))
    out: dict[str, list[list[Any]]] = {}
    for sname in book.sheet_names():
        sheet = book.sheet_by_name(sname)
        if cap and sheet.nrows > cap:
            raise ValueError(
                f"La hoja '{sname}' supera el máximo de {cap} filas. "
                f"El archivo parece demasiado grande o tener formato incorrecto."
            )
        rows: list[list[Any]] = []
        for r in range(sheet.nrows):
            row: list[Any] = []
            for c in range(sheet.ncols):
                ctype = sheet.cell_type(r, c)
                value = sheet.cell_value(r, c)
                if ctype == xlrd.XL_CELL_DATE:
                    try:
                        t = xldate_as_tuple(value, book.datemode)
                        if t[0] == 0 and t[1] == 0 and t[2] == 0:
                            # Solo hora (sin fecha) → time
                            value = time(t[3], t[4], t[5])
                        else:
                            value = datetime(*t)
                    except Exception:
                        pass
                elif ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(value)
                elif ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                row.append(value)
            rows.append(row)
        out[sname] = rows
    return out


def load_excel(path: str | Path) -> dict[str, list[list[Any]]]:
    """Universal loader: detecta xls/xlsx por magic bytes y devuelve filas como list-of-lists."""
    p = Path(path)
    if _is_xls(p):
        return _load_xls(p)
    return _load_xlsx(p)
