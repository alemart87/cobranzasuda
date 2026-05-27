"""Parse 'Bsse de llamadas' sheet from Reporte Cobranzas.

Cada fila es una llamada con: usuario, fecha, duración, dirección.
La duración viene como string `HH:MM:SS.mmm` o como time/timedelta de Excel.
"""
from __future__ import annotations

import re
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import openpyxl


SHEET_CANDIDATES = ["Bsse de llamadas", "Base de llamadas", "Bsse_de_llamadas"]
DUR_RE = re.compile(r"(\d+):(\d+):(\d+(?:\.\d+)?)")


def _parse_duration(value: Any) -> float:
    """Returns duration in seconds (float)."""
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        # Excel almacena duración como fracción de día
        return float(value) * 86400.0
    if isinstance(value, timedelta):
        return value.total_seconds()
    if isinstance(value, time):
        return value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1e6
    if isinstance(value, datetime):
        return value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1e6
    text = str(value).strip()
    m = DUR_RE.search(text)
    if m:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + float(m.group(3))
    return 0.0


def _parse_fecha(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time(0, 0))
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def parse_llamadas(path: str | Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = next((s for s in SHEET_CANDIDATES if s in wb.sheetnames), None)
    if sheet_name is None:
        # buscar cualquier hoja que tenga >100 filas y columna "Usuarios" o similar
        for s in wb.sheetnames:
            ws = wb[s]
            if ws.max_row > 100:
                headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
                if any("usuario" in h.lower() or "operador" in h.lower() for h in headers):
                    sheet_name = s
                    break
    if sheet_name is None:
        raise ValueError(
            f"No se encontró hoja con datos de llamadas. Hojas disponibles: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    # Encontrar columnas (los headers vienen mal codificados a veces: DuraciÃ³n, DirecciÃ³n)
    def col_idx(*keywords: str) -> int | None:
        for i, h in enumerate(headers):
            low = h.lower()
            if all(k.lower() in low for k in keywords):
                return i
        return None

    idx_user = col_idx("usuario") or col_idx("operador") or col_idx("agente")
    idx_fecha = col_idx("fecha")
    idx_dur = col_idx("duraci")
    idx_dir = col_idx("direcci")
    idx_cola = col_idx("cola")
    idx_concl = col_idx("conclu")

    if idx_user is None or idx_fecha is None or idx_dur is None:
        raise ValueError(
            f"Faltan columnas requeridas (Usuario, Fecha, Duración). Headers: {headers}"
        )

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        usuario = ws.cell(r, idx_user + 1).value
        if not usuario:
            continue
        fecha = _parse_fecha(ws.cell(r, idx_fecha + 1).value)
        dur_sec = _parse_duration(ws.cell(r, idx_dur + 1).value)
        rows.append({
            "usuario": str(usuario).strip(),
            "fecha": fecha,
            "duracion_seg": dur_sec,
            "direccion": (str(ws.cell(r, idx_dir + 1).value).strip() if idx_dir is not None else "Saliente"),
            "cola": (str(ws.cell(r, idx_cola + 1).value).strip() if idx_cola is not None and ws.cell(r, idx_cola + 1).value else ""),
            "conclusion": (str(ws.cell(r, idx_concl + 1).value).strip() if idx_concl is not None and ws.cell(r, idx_concl + 1).value else ""),
        })

    return rows
