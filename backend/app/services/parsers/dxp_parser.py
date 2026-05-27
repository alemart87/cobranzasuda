"""Parse DXP Voicenter XLSX into a list of dicts."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


REQUIRED_COLUMNS = [
    "Asegurado",
    "Sec.",
    "Pol.",
    "Saldo Total",
    "Hasta 30 Días",
    "Hasta 60 Días",
    "Hasta 90 Días",
    "Hasta 120 Días",
    "Hasta 150 Días",
    "Más 150 Días",
    "Organizador",
    "Nombre Sección",
]


def parse_dxp(path: str | Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Report" in wb.sheetnames:
        ws = wb["Report"]
    else:
        ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(f"DXP: faltan columnas requeridas: {missing}")

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if row.get("Asegurado"):
            rows.append(row)
    return rows
